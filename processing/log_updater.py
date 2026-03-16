from datetime import datetime, time, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment
from processing.html_report import _extract_stage_timings

TIME_FORMAT = "hh:mm:ss.000"


def _td_to_time(td):
    """Convert timedelta to datetime.time for Excel storage."""
    if td is None:
        return None
    total = td.total_seconds()
    if total < 0:
        total = 0
    h, rem = divmod(int(total), 3600)
    m, s = divmod(rem, 60)
    micro = int((total - int(total)) * 1000000)
    if h > 23:
        h = 23
        m = 59
        s = 59
        micro = 999000
    return time(h, m, s, micro)


def _fmt_duration_str(td):
    """Format timedelta as string for RUN LOG sheet."""
    if td is None:
        return ""
    total = td.total_seconds()
    sign = "-" if total < 0 else ""
    total = abs(total)
    h, rem = divmod(int(total), 3600)
    m, s = divmod(rem, 60)
    micro = int((total - int(total)) * 1000000)
    return f"{sign}{h:02d}:{m:02d}:{s:02d}.{micro:06d}"


def _next_sl_no(ws):
    last = 0
    for row in range(4, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value
        if v is not None and str(v).strip().isdigit():
            last = int(v)
    return last + 1


def _next_empty_row(ws):
    for row in range(4, ws.max_row + 2):
        if ws.cell(row=row, column=6).value is None:
            return row
    return ws.max_row + 1


def _get_stage_td(run, stage_label):
    st = _extract_stage_timings(run)
    val = st.get(stage_label)
    if isinstance(val, timedelta):
        return val
    return None


def _write_time_cell(ws, row, col, td):
    """Write a timedelta as datetime.time with hh:mm:ss.000 format."""
    cell = ws.cell(row=row, column=col)
    t = _td_to_time(td)
    if t is not None:
        cell.value = t
        cell.number_format = TIME_FORMAT


def update_run_log(historical_file, all_runs, test_name=None):
    wb = openpyxl.load_workbook(historical_file)

    grouped = defaultdict(list)
    for run in all_runs:
        grouped[run["case_count"]].append(run)

    today = datetime.now().strftime("%Y-%m-%d")
    if test_name is None:
        test_name = f"Run_{datetime.now().strftime('%d%b%Y')}"

    # === 1. Update RUN LOG sheet ===
    ws_log = wb["RUN LOG"]
    sl_no = _next_sl_no(ws_log)
    insert_row = _next_empty_row(ws_log)

    first_guid = True
    for run in all_runs:
        total_td = run["total_time"]
        wj1 = _get_stage_td(run, "WEB JOB 1 Processing Time")
        wj2 = _get_stage_td(run, "WEB JOB 2 Processing Time")
        wj3 = _get_stage_td(run, "WEB JOB 3 Processing Time")
        cc = run["case_count"]
        tpc = total_td.total_seconds() / cc if cc > 0 else 0

        if first_guid:
            ws_log.cell(row=insert_row, column=2, value=sl_no)
            ws_log.cell(row=insert_row, column=3, value=today)
            ws_log.cell(row=insert_row, column=4, value="AUTOMATED")
            ws_log.cell(row=insert_row, column=5, value=test_name)
            first_guid = False

        ws_log.cell(row=insert_row, column=6, value=_fmt_duration_str(total_td))
        ws_log.cell(row=insert_row, column=7, value=_fmt_duration_str(wj1) if wj1 else "NA")
        ws_log.cell(row=insert_row, column=8, value=_fmt_duration_str(wj2) if wj2 else "NA")
        ws_log.cell(row=insert_row, column=9, value=_fmt_duration_str(wj3) if wj3 else "NA")
        ws_log.cell(row=insert_row, column=10, value="NA")
        ws_log.cell(row=insert_row, column=11, value=round(tpc, 3))
        insert_row += 1

    # === 2. Add detail sheet (matching R18.0.1 format exactly) ===
    sheet_name = f"Auto_{datetime.now().strftime('%d%b%Y_%H%M')}"
    sheet_name = sheet_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for cc, runs in sorted(grouped.items()):
        total_rec = cc * len(runs)
        num_runs = len(runs)

        # Row 1: empty
        # Row 2: FILES header + per-GUID columns
        row = 2
        cell = ws.cell(row=row, column=1, value=f"FILES ({total_rec} REC)")
        cell.font = bold
        for i, run in enumerate(runs):
            guid = run["job_guid"]
            header = f"JOB GUID - {guid}\n({cc} REC)"
            cell = ws.cell(row=row, column=2 + i, value=header)
            cell.alignment = wrap

        # Row 3: STATUS
        row = 3
        ws.cell(row=row, column=1, value="STATUS").font = bold
        for i in range(num_runs):
            ws.cell(row=row, column=2 + i, value="Job successfully completed")

        # Row 4: CREATE JOB >>
        row = 4
        ws.cell(row=row, column=1, value="CREATE JOB  >>").font = bold
        ws.cell(row=row, column=2, value="TRIGGER")

        # Stage timing rows (5-10)
        stage_rows = [
            (5, "WAIT TIME 0 (START - WJ1) -- >", "WAIT TIME 0 (START - WJ1)"),
            (6, "WEB JOB 1 Processing time", "WEB JOB 1 Processing Time"),
            (7, "WAIT TIME 1 (WJ2 - WJ1) -- >", "WAIT TIME 1 (WJ2 - WJ1)"),
            (8, "WEB JOB 2 Processing time", "WEB JOB 2 Processing Time"),
            (9, "WAIT TIME 2 (WJ3 - WJ2) -- >", "WAIT TIME 2 (WJ3 - WJ2)"),
            (10, "WEB JOB 3 Processing time", "WEB JOB 3 Processing Time"),
        ]

        for row, sheet_label, key in stage_rows:
            ws.cell(row=row, column=1, value=sheet_label).font = bold
            for i, run in enumerate(runs):
                td = _get_stage_td(run, key)
                _write_time_cell(ws, row, 2 + i, td)

        # Row 11: WAIT TIME 3 (placeholder)
        ws.cell(row=11, column=1, value="WAIT TIME 3 (WJ4 - WJ3) -- >").font = bold

        # Row 12: RETRY JOB (placeholder)
        ws.cell(row=12, column=1, value="RETRY JOB Processing time").font = bold

        # Row 13: TOTAL PROCESSING TIME
        ws.cell(row=13, column=1, value="TOTAL PROCESSING TIME \n(hh:mm:ss.000)").font = bold
        ws.cell(row=13, column=1).alignment = wrap
        for i, run in enumerate(runs):
            _write_time_cell(ws, 13, 2 + i, run["total_time"])

        # Row 14: AVG. Time per input case
        ws.cell(row=14, column=1, value="AVG. Time per input case").font = bold
        for i, run in enumerate(runs):
            cc_val = run["case_count"]
            if cc_val > 0:
                tpc_td = run["total_time"] / cc_val
                _write_time_cell(ws, 14, 2 + i, tpc_td)

    # Auto-fit column A width
    ws.column_dimensions["A"].width = 38

    wb.save(historical_file)
    print(f"GDC_RUN_LOG.xlsx updated: new rows in 'RUN LOG' + new sheet '{sheet_name}'")
    return sheet_name
